# src/validation_agent/workbook_loader.py
from __future__ import annotations
from pathlib import Path
from typing import List
import textwrap
import zipfile
import xml.etree.ElementTree as ET
import json
import subprocess
import tempfile
import os
import logging
import sys
logger = logging.getLogger(__name__)
EXCEL_SUFFIXES = {".xlsm", ".xls", ".xlsx", ".xlsb"}
def _run_pbitools_extract(pbix_path: Path) -> Path | None:
    exe = os.getenv("PBI_TOOLS_EXE")
    if not exe and getattr(sys, "frozen", False):
        exe_candidate = Path(sys.executable).with_name("pbi-tools.exe")
        if exe_candidate.exists():
            exe = str(exe_candidate)
    if not exe:
        exe = "pbi-tools"
    tmp_dir = tempfile.mkdtemp(prefix="pbitools_extract_")
    out_dir = Path(tmp_dir)
    cmd = [
        exe,
        "extract",
        str(pbix_path),
        "-extractFolder", str(out_dir),
        "-modelSerialization", "Legacy",      # ensures Model/DataModelSchema.json
        "-mashupSerialization", "Default",
    ]
    logger.info("pbi-tools: starting extract")
    logger.info("pbi-tools: exe=%r tmp_dir=%s", exe, out_dir)
    logger.info("pbi-tools: full command: %s", " ".join(cmd))
    try:
        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
    except FileNotFoundError:
        logger.error(
            "pbi-tools executable %r not found. "
            "Make sure it is installed in this environment and on PATH, "
            "or set PBI_TOOLS_EXE to an absolute path.",
            exe,
        )
        return None
    except Exception as exc:
        logger.exception("Unexpected error when running pbi-tools: %s", exc)
        return None
    logger.info("pbi-tools: return code = %s", result.returncode)
    if result.stdout:
        logger.info("pbi-tools STDOUT:\n%s", result.stdout)
    if result.stderr:
        logger.info("pbi-tools STDERR:\n%s", result.stderr)
    if result.returncode != 0:
        logger.error("pbi-tools extract failed with code %s; skipping JSON model.", result.returncode)
        return None
    logger.info("pbi-tools extract succeeded, output directory: %s", out_dir)
    return out_dir
def _pbitools_model_summary(extract_root: Path, max_tables: int = 40) -> str:
    model_dir = extract_root / "Model"
    candidates: list[Path] = []
    for name in ("DataModelSchema.json", "DataModelSchema"):
        p = model_dir / name
        if p.exists():
            candidates.append(p)
    if model_dir.exists():
        for p in model_dir.iterdir():
            if "datamodelschema" in p.name.lower() and p not in candidates:
                candidates.append(p)
    db_path = model_dir / "database.json"
    if db_path.exists() and db_path not in candidates:
        candidates.append(db_path)
    schema_path: Path | None = None
    for p in candidates:
        if p.exists():
            schema_path = p
            break
    if not schema_path:
        logger.warning(
            "pbi-tools: no model schema found under %s (tried DataModelSchema*, database.json)",
            model_dir,
        )
        return ""
    logger.info("pbi-tools: looking for model schema at %s", schema_path)
    try:
        raw = schema_path.read_text(encoding="utf-8")
        data = json.loads(raw)
    except Exception as exc:
        logger.exception("pbi-tools: failed to parse %s: %s", schema_path, exc)
        return ""
    if "database" in data:
        model = (data.get("database") or {}).get("model") or {}
    else:
        model = data.get("model") or data
    def _extract_embedded_model_from_annotations(m: dict) -> dict | None:
        anns = m.get("annotations") or []
        if not isinstance(anns, list):
            return None
        for ann in anns:
            if not isinstance(ann, dict):
                continue
            val = ann.get("value")
            if not isinstance(val, str):
                continue
            if '"tables"' not in val and '"model"' not in val:
                continue
            try:
                inner = json.loads(val)
            except Exception:
                continue
            if not isinstance(inner, dict):
                continue
            if "database" in inner:
                inner_model = (inner.get("database") or {}).get("model") or {}
            else:
                inner_model = inner.get("model") or inner
            if isinstance(inner_model, dict) and (
                inner_model.get("tables") or inner_model.get("relationships")
            ):
                logger.info(
                    "pbi-tools: found embedded model in annotations; keys=%r",
                    sorted(inner_model.keys()),
                )
                return inner_model
        return None
    if not (model.get("tables") or []):
        embedded = _extract_embedded_model_from_annotations(model)
        if embedded:
            model = embedded
    logger.info("pbi-tools: model top-level keys: %r", list(model.keys()))
    tables = model.get("tables") or []
    relationships = model.get("relationships") or []
    if not tables:
        tables_dir = model_dir / "tables"
        synthesised: list[dict] = []
        if tables_dir.exists():
            logger.info("pbi-tools: no tables in JSON; scanning %s for table folders", tables_dir)
            for tdir in sorted(tables_dir.iterdir()):
                if not tdir.is_dir():
                    continue
                tname = tdir.name
                columns: list[dict] = []
                measures: list[dict] = []
                table_meta = tdir / "table.json"
                if table_meta.exists():
                    try:
                        meta_raw = table_meta.read_text(encoding="utf-8")
                        meta = json.loads(meta_raw)
                        for col in meta.get("columns", []) or []:
                            cname = col.get("name")
                            if cname:
                                columns.append({"name": cname})
                    except Exception as exc:
                        logger.debug("pbi-tools: failed to parse %s: %s", table_meta, exc)
                cols_dir = tdir / "columns"
                if cols_dir.exists():
                    for col_file in cols_dir.iterdir():
                        if col_file.suffix.lower() not in (".json", ".dax"):
                            continue
                        cname = col_file.stem
                        if cname and not any(c["name"] == cname for c in columns):
                            columns.append({"name": cname})
                measures_dir = tdir / "measures"
                if measures_dir.exists():
                    for j in measures_dir.glob("*.json"):
                        mname = j.stem
                        expr = ""
                        dax_file = measures_dir / f"{mname}.dax"
                        if dax_file.exists():
                            try:
                                expr = dax_file.read_text(encoding="utf-8", errors="ignore").strip()
                            except Exception:
                                expr = ""
                        if not expr:
                            try:
                                m_meta = json.loads(j.read_text(encoding="utf-8"))
                                expr = (m_meta.get("expression") or "").strip()
                            except Exception:
                                expr = ""
                        measures.append({"name": mname, "expression": expr})
                if columns or measures:
                    synthesised.append(
                        {
                            "name": tname,
                            "columns": columns,
                            "measures": measures,
                        }
                    )
            if synthesised:
                logger.info(
                    "pbi-tools: synthesised %d tables from Model/tables folder",
                    len(synthesised),
                )
                tables = synthesised
    def _walk(obj):
        if isinstance(obj, dict):
            for v in obj.values():
                yield from _walk(v)
        elif isinstance(obj, list):
            for item in obj:
                yield from _walk(item)
    def _find_tables_fallback(m) -> list[dict]:
        best: list[dict] = []
        for node in _walk(m):
            if not isinstance(node, list) or not node:
                continue
            if not all(isinstance(x, dict) for x in node):
                continue
            if any("name" in t for t in node) and any(
                any(k in t for k in ("columns", "measures", "partitions", "hierarchies"))
                for t in node
            ):
                best = node
                break
        return best
    def _find_relationships_fallback(m) -> list[dict]:
        for node in _walk(m):
            if not isinstance(node, list) or not node or not isinstance(node[0], dict):
                continue
            if any(
                (("fromTable" in r or "fromTableName" in r)
                 and ("toTable" in r or "toTableName" in r))
                for r in node
            ):
                return node
        return []
    if not tables:
        fallback_tables = _find_tables_fallback(model)
        if fallback_tables:
            tables = fallback_tables
            logger.info(
                "pbi-tools: recovered %d tables via heuristic search",
                len(tables),
            )
    if not relationships:
        fallback_rels = _find_relationships_fallback(model)
        if fallback_rels:
            relationships = fallback_rels
            logger.info(
                "pbi-tools: recovered %d relationships via heuristic search",
                len(relationships),
            )
    logger.info(
        "pbi-tools: model has %d tables, %d relationships (after heuristics)",
        len(tables),
        len(relationships),
    )
    if not tables and not relationships:
        return ""
    parts: List[str] = ["## Data model (from pbi-tools)"]
    if tables:
        def _sort_key(tbl: dict) -> tuple[int, str]:
            name = (tbl.get("name") or "").lower()
            if name in {"f_center", "f_coding"}:
                return (0, name)
            if name.startswith("f_"):
                return (1, name)
            if name.startswith("dim_") or name.startswith("d_"):
                return (2, name)
            return (3, name)
        tables_sorted = sorted(tables, key=_sort_key)
        for tbl in tables_sorted:
            tname = tbl.get("name") or "<unnamed table>"
            cols = [c.get("name") for c in tbl.get("columns", []) if c.get("name")]
            measures = tbl.get("measures") or []
            calc_cols = [
                c
                for c in (tbl.get("columns") or [])
                if (c.get("expression") or "").strip()
            ]
            logger.info(
                "pbi-tools: table %r, cols=%d, measures=%d, calc_cols=%d",
                tname,
                len(cols),
                len(measures),
                len(calc_cols),
            )
            parts.append(f"\n### Table: {tname}")
            if cols:
                col_list = ", ".join(cols)   # ALL columns
                parts.append(f"- Columns: {col_list}")
            if measures:
                parts.append("- Measures:")
                for m in measures:           # ALL measures
                    mname = m.get("name") or "<unnamed measure>"
                    expr = (m.get("expression") or "").strip()
                    if expr:
                        parts.append(f"  - {mname}: {expr}")   # full DAX
                    else:
                        parts.append(f"  - {mname}")
            if calc_cols:
                parts.append("- Calculated columns:")
                for c in calc_cols:          # ALL calc cols
                    cname = c.get("name") or "<unnamed column>"
                    expr = (c.get("expression") or "").strip()
                    if expr:
                        parts.append(f"  - {cname}: {expr}")
                    else:
                        parts.append(f"  - {cname}")
    if relationships:
        parts.append("\n## Relationships")
        for rel in relationships:
            from_tbl = rel.get("fromTable") or rel.get("fromTableName") or "?"
            from_col = rel.get("fromColumn") or rel.get("fromColumnName") or "?"
            to_tbl = rel.get("toTable") or rel.get("toTableName") or "?"
            to_col = rel.get("toColumn") or rel.get("toColumnName") or "?"
            is_active = rel.get("isActive")
            active_flag = "" if is_active in (None, True) else " (inactive)"
            parts.append(
                f"- {from_tbl}[{from_col}] → {to_tbl}[{to_col}]{active_flag}"
            )
    return "\n".join(parts).strip()
def _pbitools_layout_summary(extract_root: Path, max_visuals_per_page: int = 10) -> str:
    layout_path = extract_root / "Report" / "Layout"
    logger.info("pbi-tools: looking for layout at %s", layout_path)
    if not layout_path.exists():
        logger.warning("pbi-tools: Layout file not found, skipping layout summary.")
        return ""
    raw = layout_path.read_bytes()
    try:
        text = raw.decode("utf-16le")
    except UnicodeDecodeError:
        logger.info("pbi-tools: Layout not UTF-16LE, falling back to UTF-8.")
        text = raw.decode("utf-8", errors="ignore")
    try:
        data = json.loads(text)
    except Exception as exc:
        logger.exception("pbi-tools: failed to parse Layout JSON: %s", exc)
        return ""
    sections = data.get("sections", []) or []
    logger.info("pbi-tools: layout has %d sections/pages", len(sections))
    parts: List[str] = ["## Report pages and visuals (from pbi-tools)"]
    for section in sections:
        title = section.get("displayName") or section.get("name") or "<untitled page>"
        visuals = section.get("visualContainers") or []
        logger.info("pbi-tools: page %r has %d visuals", title, len(visuals))
        parts.append(f"\n### Page: {title}")
        parts.append(f"- Visual count: {len(visuals)}")
        for vc in visuals[:max_visuals_per_page]:
            cfg_raw = vc.get("config")
            vis_title = None
            vis_type = None
            try:
                cfg = json.loads(cfg_raw) if isinstance(cfg_raw, str) else cfg_raw
            except Exception:
                cfg = None
            if cfg:
                single = cfg.get("singleVisual") or cfg.get("singleVisualGroup") or {}
                title_obj = single.get("title") or {}
                vis_title = title_obj.get("text") or single.get("displayName")
                vis_type = single.get("visualType") or single.get("groupType")
            parts.append(f"  - {vis_title or '<untitled visual>'} ({vis_type or 'visual'})")
        if len(visuals) > max_visuals_per_page:
            parts.append(f"  - ... {len(visuals) - max_visuals_per_page} more visuals not listed")
    return "\n".join(parts).strip()
def extract_excel_context(path: Path, max_chars: int = 0) -> str:
    suffix = path.suffix.lower()
    if suffix not in EXCEL_SUFFIXES:
        raise ValueError(f"Not an Excel workbook: {path}")
    parts: List[str] = [f"# Excel workbook: {path.name}"]
    structure = _extract_excel_structure(path)
    if structure:
        parts.append(structure)
    sql_section = _extract_excel_sql_queries(path)
    if sql_section:
        parts.append(sql_section)
    com_section = _extract_excel_com_details(
        path,
        max_queries=0,   # 0 == no per-section cap
        max_chars=0,     # 0 == no truncation inside
    )
    if com_section:
        parts.append(com_section)
    vba_section = _extract_excel_vba(path, max_chars=0)
    if vba_section:
        parts.append(vba_section)
    text = "\n".join(part for part in parts if part).strip()
    return text if not max_chars else text[:max_chars]
def _extract_excel_structure(path: Path) -> str:
    parts: List[str] = []
    suffix = path.suffix.lower()
    if suffix == ".xlsb":
        try:
            from pyxlsb import open_workbook  # type: ignore
        except Exception:
            return (
                "[Workbook structure for .xlsb not available: "
                "install 'pyxlsb' to list sheets and basic structure.]"
            )
        try:
            with open_workbook(str(path)) as wb:
                parts.append("## Sheets")
                for sheet in wb.sheets:
                    try:
                        name = getattr(sheet, "name", None) or str(sheet)
                    except Exception:
                        name = str(sheet)
                    parts.append(f"- {name}")
        except Exception as exc:
            return (
                f"[Could not read .xlsb workbook structure via pyxlsb: "
                f"{type(exc).__name__}: {exc}]"
            )
        return "\n".join(parts).strip()
    try:
        from openpyxl import load_workbook  # type: ignore
        wb = load_workbook(filename=str(path), data_only=False, keep_links=True)
        sheet_names = list(wb.sheetnames)
        parts.append("## Sheets")
        for s in sheet_names:
            parts.append(f"- {s}")
        try:
            defined = getattr(wb, "defined_names", None)
            if defined and getattr(defined, "definedName", None):
                parts.append("\n## Named ranges")
                for dn in defined.definedName:
                    parts.append(f"- {dn.name}: {dn.attr_text}")
        except Exception:
            pass
    except Exception:
        parts.append("\n[Could not read workbook structure via openpyxl.]")
    return "\n".join(parts).strip()
def _extract_excel_sql_queries(path: Path, max_queries: int = 0) -> str:
    queries: List[str] = []
    connections: List[str] = []
    suffix = path.suffix.lower()
    try:
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
            if "xl/connections.xml" in names:
                try:
                    xml_bytes = zf.read("xl/connections.xml")
                    root = ET.fromstring(xml_bytes)
                    for node in root.iter():
                        tag = node.tag.lower()
                        if not tag.endswith("dbpr"):
                            continue
                        cmd = node.attrib.get("command", "") or ""
                        conn = node.attrib.get("connection", "") or ""
                        if cmd.strip():
                            queries.append(cmd.strip())
                        if conn.strip():
                            connections.append(conn.strip())
                except Exception:
                    pass
            for name in names:
                if not (
                    name.startswith("xl/queryTables/")
                    and name.lower().endswith(".xml")
                ):
                    continue
                try:
                    xml_bytes = zf.read(name)
                    root = ET.fromstring(xml_bytes)
                    for node in root.iter():
                        tag = node.tag.lower()
                        if tag.endswith("command") or tag.endswith("sql"):
                            txt = (node.text or "").strip()
                            if txt:
                                queries.append(txt)
                except Exception:
                    continue
            m_queries: List[str] = []
            for name in names:
                if not (name.startswith("xl/queries/") and name.lower().endswith(".xml")):
                    continue
                try:
                    xml_bytes = zf.read(name)
                    root = ET.fromstring(xml_bytes)
                except Exception:
                    continue
                for node in root.iter():
                    tag = node.tag.lower()
                    if tag.endswith("m"):
                        txt = (node.text or "").strip()
                        if txt:
                            m_queries.append(txt)
                            if max_queries and len(m_queries) >= max_queries:
                                break
                if len(m_queries) >= max_queries:
                    break
            if not queries:
                exts = (".xml", ".rels", ".txt")
                if suffix == ".xlsb":
                    exts = exts + (".bin",)
                for name in names:
                    lname = name.lower()
                    if not lname.endswith(exts):
                        continue
                    try:
                        raw = zf.read(name)
                    except Exception:
                        continue
                    try:
                        text = raw.decode("utf-8", errors="ignore")
                    except Exception:
                        text = raw.decode("latin-1", errors="ignore")
                    lowered = text.lower()
                    if "select " not in lowered or " from " not in lowered:
                        continue
                    for line in text.splitlines():
                        lline = line.lower()
                        if "select " in lline and " from " in lline:
                            stripped = line.strip()
                            if stripped:
                                queries.append(stripped)
                                if len(queries) >= max_queries:
                                    break
                    if len(queries) >= max_queries:
                        break
    except Exception:
        return ""
    if not queries and not connections and not m_queries:
        return ""
    section: List[str] = ["\n## SQL / Power Query / connections"]
    if connections:
        section.append("\n### Connections")
        for idx, conn in enumerate(connections, start=1):
            section.append(
                f"- Connection {idx}: "
                f"{textwrap.shorten(conn, width=200, placeholder='...')}"
            )
    if queries:
        section.append("\n### SQL Queries")
        for idx, q in enumerate(queries, start=1):
            section.append(
                f"\n#### Query {idx}\n"
                "```sql\n"
                f"{q}\n"
                "```"
            )
    if m_queries:
        section.append("\n### Power Query (M) scripts")
        for idx, m in enumerate(m_queries, start=1):
            section.append(
                f"\n#### M Query {idx}\n"
                "```m\n"
                f"{m}\n"
                "```"
            )
    return "\n".join(section).strip()
def _extract_excel_com_details(
    path: Path,
    max_queries: int = 0,   # 0 = no limit
    max_chars: int = 0,      # 0 = no truncation
) -> str:
    try:
        import win32com.client  # type: ignore
    except Exception:
        logger.info("COM Excel not available (win32com import failed); skipping COM extraction.")
        return ""
    try:
        excel = win32com.client.DispatchEx("Excel.Application")  # type: ignore[attr-defined]
    except Exception as exc:
        logger.info("COM Excel DispatchEx failed: %s; skipping COM extraction.", exc)
        return ""
    try:
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(str(path), ReadOnly=True)
        except Exception as exc:
            logger.exception("COM Excel failed to open workbook %s: %s", path, exc)
            try:
                excel.Quit()
            except Exception:
                pass
            return ""
        try:
            parts: List[str] = ["\n## COM-based workbook inspection"]
            try:
                sheet_names: list[str] = []
                for sh in wb.Worksheets:
                    try:
                        sheet_names.append(str(sh.Name))
                    except Exception:
                        continue
                if sheet_names:
                    parts.append("\n### Sheets (COM)")
                    for nm in sheet_names:
                        parts.append(f"- {nm}")
            except Exception:
                logger.debug("COM: failed to enumerate sheets", exc_info=True)
            try:
                if wb.Names.Count:
                    parts.append("\n### Named ranges (COM)")
                    count = 0
                    for name_obj in wb.Names:
                        try:
                            nm = str(name_obj.Name)
                            refers = str(name_obj.RefersTo)
                        except Exception:
                            continue
                        parts.append(f"- {nm}: {refers}")
            except Exception:
                logger.debug("COM: failed to enumerate named ranges", exc_info=True)
            try:
                table_lines: list[str] = []
                for sh in wb.Worksheets:
                    try:
                        lo_collection = getattr(sh, "ListObjects", None)
                    except Exception:
                        lo_collection = None
                    if not lo_collection:
                        continue
                    for lo in lo_collection:
                        try:
                            tname = str(lo.Name)
                        except Exception:
                            tname = "<unnamed>"
                        try:
                            rng = lo.Range
                            addr = str(rng.Address(False, False))
                        except Exception:
                            addr = "?"
                        try:
                            sheet_name = str(sh.Name)
                        except Exception:
                            sheet_name = "?"
                        conn_name = ""
                        try:
                            qt = getattr(lo, "QueryTable", None)
                            if qt is not None:
                                try:
                                    wb_conn = getattr(qt, "WorkbookConnection", None)
                                    if wb_conn is not None:
                                        conn_name = str(getattr(wb_conn, "Name", "") or "")
                                except Exception:
                                    conn_name = str(getattr(qt, "Connection", "") or "")
                        except Exception:
                            pass
                        line = f"- {tname} (sheet: {sheet_name}, range: {addr}"
                        if conn_name:
                            line += f", connection: {conn_name}"
                        line += ")"
                        table_lines.append(line)
                if table_lines:
                    parts.append("\n### Tables / ListObjects (COM)")
                    parts.extend(table_lines if not max_queries else table_lines[:max_queries])
            except Exception:
                logger.debug("COM: failed to enumerate ListObjects", exc_info=True)
            try:
                conn_lines: list[str] = []
                for conn in wb.Connections:
                    try:
                        cname = str(conn.Name)
                    except Exception:
                        cname = "<unnamed>"
                    conn_str = ""
                    cmd_text = ""
                    try:
                        if hasattr(conn, "OLEDBConnection"):
                            oledb = conn.OLEDBConnection
                            conn_str = str(getattr(oledb, "Connection", "") or "")
                            cmd_text = str(getattr(oledb, "CommandText", "") or "")
                        elif hasattr(conn, "ODBCConnection"):
                            odbc = conn.ODBCConnection
                            conn_str = str(getattr(odbc, "Connection", "") or "")
                            cmd_text = str(getattr(odbc, "CommandText", "") or "")
                    except Exception:
                        pass
                    if not conn_str:
                        try:
                            conn_str = str(getattr(conn, "Description", "") or "")
                        except Exception:
                            pass
                    conn_str_short = conn_str
                    cmd_text_short = str(cmd_text)
                    line = f"- {cname}"
                    if conn_str_short:
                        line += f" | connection: {conn_str_short}"
                    if cmd_text_short:
                        line += f" | command: {cmd_text_short}"
                    conn_lines.append(line)
                if conn_lines:
                    parts.append("\n### Connections (COM)")
                    parts.extend(conn_lines if not max_queries else conn_lines[:max_queries])
            except Exception:
                logger.debug("COM: failed to enumerate Connections", exc_info=True)
            try:
                qt_lines: list[str] = []
                for sh in wb.Worksheets:
                    try:
                        qts = getattr(sh, "QueryTables", None)
                    except Exception:
                        qts = None
                    if not qts:
                        continue
                    for qt in qts:
                        try:
                            qname = str(qt.Name)
                        except Exception:
                            qname = "<unnamed>"
                        try:
                            cmd = qt.CommandText
                        except Exception:
                            cmd = ""
                        try:
                            conn = qt.Connection
                        except Exception:
                            conn = ""
                        cmd_short = str(cmd)
                        conn_short = str(conn)
                        line = f"- {qname} (sheet: {sh.Name})"
                        if conn_short:
                            line += f" | connection: {conn_short}"
                        if cmd_short:
                            line += f" | command: {cmd_short}"
                        qt_lines.append(line)
                if qt_lines:
                    parts.append("\n### QueryTables (COM)")
                    parts.extend(qt_lines if not max_queries else qt_lines[:max_queries])
            except Exception:
                logger.debug("COM: failed to enumerate QueryTables", exc_info=True)
            try:
                m_parts: list[str] = []
                queries_col = getattr(wb, "Queries", None)
                if queries_col is not None:
                    count = 0
                    for q in queries_col:
                        try:
                            qname = str(q.Name)
                            formula = str(q.Formula)
                        except Exception:
                            continue
                        formula_short = formula
                        m_parts.append(
                            f"\n#### M Query (COM): {qname}\n"
                            "```m\n"
                            f"{formula_short}\n"
                            "```"
                        )
                        count += 1
                        if max_queries and count >= max_queries:
                            break
                if m_parts:
                    parts.append("\n### Power Query (M) via COM")
                    parts.extend(m_parts)
            except Exception:
                logger.debug("COM: failed to enumerate Workbook.Queries", exc_info=True)
            try:
                pivot_lines: list[str] = []
                for sh in wb.Worksheets:
                    try:
                        pivots = getattr(sh, "PivotTables", None)
                    except Exception:
                        pivots = None
                    if not pivots:
                        continue
                    for pt in pivots:
                        try:
                            pname = str(pt.Name)
                        except Exception:
                            pname = "<unnamed>"
                        try:
                            src_name = str(getattr(pt, "SourceData", "") or "")
                        except Exception:
                            src_name = ""
                        src_short = src_name
                        line = f"- {pname} (sheet: {sh.Name}"
                        if src_short:
                            line += f", source: {src_short}"
                        line += ")"
                        pivot_lines.append(line)
                if pivot_lines:
                    parts.append("\n### PivotTables (COM)")
                    parts.extend(pivot_lines if not max_queries else pivot_lines[:max_queries])
            except Exception:
                logger.debug("COM: failed to enumerate PivotTables", exc_info=True)
            text = "\n".join(parts).strip()
            return text if not max_chars else text[:max_chars]
        finally:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                excel.Quit()
            except Exception:
                pass
    except Exception:
        logger.exception("COM-based Excel extraction failed for %s", path)
        return ""
def _extract_excel_vba(path: Path, max_chars: int = 0) -> str:
    parts: List[str] = []
    try:
        from oletools.olevba import VBA_Parser  # type: ignore
        vba = VBA_Parser(str(path))
        try:
            if vba.detect_vba_macros():
                parts.append("\n## VBA macros")
                for (_, _, vba_filename, vba_code) in vba.extract_all_macros():
                    if not vba_code:
                        continue
                    snippet = vba_code  # full module
                    parts.append(
                        f"\n### Module: {vba_filename}\n"
                        "```vba\n"
                        f"{snippet}\n"
                        "```"
                    )
        finally:
            vba.close()
    except Exception:
        parts.append("\n[No VBA macros extracted or oletools not installed.]")
    text = "\n".join(parts).strip()
    return text if not max_chars else text[:max_chars]
def _pbix_extract_layout(zf: zipfile.ZipFile, names: List[str], max_visuals_per_page: int = 10) -> str:
    if "Report/Layout" not in names:
        return ""
    try:
        raw = zf.read("Report/Layout")
    except KeyError:
        return ""
    try:
        text = raw.decode("utf-16le")
    except UnicodeDecodeError:
        text = raw.decode("utf-8", errors="ignore")
    try:
        data = json.loads(text)
    except Exception:
        return ""
    parts: List[str] = ["## Report pages and visuals (best effort)"]
    for section in data.get("sections", []):
        title = section.get("displayName") or section.get("name") or "<untitled page>"
        visuals = section.get("visualContainers") or []
        parts.append(f"\n### Page: {title}")
        parts.append(f"- Visual count: {len(visuals)}")
        for vc in visuals[:max_visuals_per_page]:
            cfg_raw = vc.get("config")
            vis_title = None
            vis_type = None
            try:
                cfg = json.loads(cfg_raw) if isinstance(cfg_raw, str) else cfg_raw
            except Exception:
                cfg = None
            if cfg:
                single = cfg.get("singleVisual") or cfg.get("singleVisualGroup") or {}
                title_obj = single.get("title") or {}
                vis_title = title_obj.get("text") or single.get("displayName")
                vis_type = single.get("visualType") or single.get("groupType")
            parts.append(f"  - {vis_title or '<untitled visual>'} ({vis_type or 'visual'})")
        if len(visuals) > max_visuals_per_page:
            parts.append(f"  - ... {len(visuals) - max_visuals_per_page} more visuals not listed")
    return "\n".join(parts).strip()
def _pbix_extract_queries(
    zf: zipfile.ZipFile,
    names: List[str],
    max_queries: int = 40,
) -> str:
    queries: List[str] = []
    for name in names:
        if not name.lower().endswith(".dax"):
            continue
        try:
            raw = zf.read(name)
            try:
                text = raw.decode("utf-16le")
            except UnicodeDecodeError:
                text = raw.decode("utf-8", errors="ignore")
        except Exception:
            continue
        stripped = text.strip()
        if not stripped:
            continue
        header = f"-- {name}"
        queries.append(f"{header}\n{stripped}")
        if len(queries) >= max_queries:
            break
    if len(queries) < max_queries:
        for name in names:
            lname = name.lower()
            if not any(token in lname for token in ("datamashup", "section", "report", ".json", ".xml", ".txt")):
                continue
            try:
                raw = zf.read(name)
                try:
                    text = raw.decode("utf-8")
                except UnicodeDecodeError:
                    text = raw.decode("utf-16le", errors="ignore")
            except Exception:
                continue
            lowered = text.lower()
            if "select " not in lowered or " from " not in lowered:
                continue
            for line in text.splitlines():
                lline = line.lower()
                if "select " in lline and " from " in lline:
                    stripped = line.strip()
                    if stripped:
                        queries.append(stripped)
                        if len(queries) >= max_queries:
                            break
            if len(queries) >= max_queries:
                break
    if not queries:
        return ""
    parts: List[str] = ["## DAX / SQL queries (best effort)"]
    for idx, q in enumerate(queries, start=1):
        shortened = textwrap.shorten(q, width=800, placeholder=" …")
        parts.append(
            f"\n### Query {idx}\n"
            "```sql\n"
            f"{shortened}\n"
            "```"
        )
    return "\n".join(parts).strip()
def extract_pbix_context(path: Path, max_chars: int = 0) -> str:
    suffix = path.suffix.lower()
    if suffix != ".pbix":
        raise ValueError(f"Not a PBIX file: {path}")
    logger.info("PBIX: starting context extraction for %s", path)
    header = f"# Power BI PBIX: {path.name}\n"
    parts: List[str] = []
    extract_root = _run_pbitools_extract(path)
    if extract_root is not None:
        logger.info("PBIX: using pbi-tools extract at %s", extract_root)
        model_section = _pbitools_model_summary(extract_root)
        if model_section:
            logger.info("PBIX: pbi-tools model summary length = %d chars", len(model_section))
            parts.append(model_section)
        layout_section = _pbitools_layout_summary(extract_root)
        if layout_section:
            logger.info("PBIX: pbi-tools layout summary length = %d chars", len(layout_section))
            parts.append(layout_section)
    else:
        logger.warning("PBIX: pbi-tools extract unavailable; falling back to raw ZIP heuristics.")
    try:
        logger.info("PBIX: opening file as zip for additional heuristic extraction.")
        with zipfile.ZipFile(path) as zf:
            names = zf.namelist()
            logger.info("PBIX: zip contains %d members", len(names))
            has_tables = any("### Table:" in p for p in parts)
            if not has_tables:
                model_section = _pbix_extract_model(zf, names)
                if model_section:
                    logger.info("PBIX: heuristic model summary length = %d chars", len(model_section))
                    parts.append(model_section)
            layout_section = _pbix_extract_layout(zf, names)
            if layout_section:
                logger.info("PBIX: heuristic layout summary length = %d chars", len(layout_section))
                parts.append(layout_section)
            if not parts:
                if any(n.endswith("DataModel") or n.lower().endswith("datamodel") for n in names):
                    msg = (
                        "[PBIX contains a binary DataModel that this helper does not decode. "
                        "Install pbi-tools and ensure it is on PATH (or set PBI_TOOLS_EXE) "
                        "to extract the Tabular model as JSON.]"
                    )
                    logger.warning("PBIX: %s", msg)
                    parts.append(msg)
                else:
                    msg = (
                        "[No SQL-like queries or model details were detected in this PBIX. "
                        "Queries may be stored in mashups or require a dedicated PBIX parser.]"
                    )
                    logger.warning("PBIX: %s", msg)
                    parts.append(msg)
    except Exception as exc:
        logger.exception("PBIX: zip-based extraction failed: %s", exc)
        parts.append(
            f"[PBIX extraction failed: {type(exc).__name__}: {exc}. "
            "To get full SQL and model details, ensure pbi-tools is installed and reachable.]"
        )
    text = header + "\n".join(parts)
    logger.info("PBIX: final extracted context length = %d chars", len(text))
    logger.info("PBIX: preview of extracted context:\n%s", text[:500])
    return text if not max_chars else text[:max_chars]
def _pbix_extract_model(zf: zipfile.ZipFile, names: List[str]) -> str:
    candidates = [
        n for n in names
        if "datamodel" in n.lower() and n.lower().endswith(".json")
    ]
    if not candidates:
        return ""
    parts: List[str] = ["## Data model (best effort)"]
    for name in candidates:
        try:
            raw = zf.read(name)
            text = raw.decode("utf-8", errors="ignore")
            data = json.loads(text)
        except Exception:
            continue
        model = data.get("model") or data
        tables = model.get("tables") or []
        for tbl in tables:
            tname = tbl.get("name")
            if not tname:
                continue
            cols = [c.get("name") for c in tbl.get("columns", []) if c.get("name")]
            measures = tbl.get("measures", []) or []
            parts.append(f"\n### Table: {tname}")
            if cols:
                col_list = ", ".join(cols[:20])
                if len(cols) > 20:
                    col_list += " …"
                parts.append(f"- Columns: {col_list}")
            if measures:
                parts.append("- Measures:")
                for m in measures:   # ALL measures
                    mname = m.get("name") or "<unnamed>"
                    expr = (m.get("expression") or "").strip()
                    if expr:
                        parts.append(f"  - {mname}: {expr}")   # full DAX
                    else:
                        parts.append(f"  - {mname}")
    return "\n".join(parts).strip() if len(parts) > 1 else ""
def _pbix_extract_sql_like(zf: zipfile.ZipFile, names: List[str]) -> str:
    queries: List[str] = []
    for name in names:
        lname = name.lower()
        if not any(token in lname for token in (
            "datamashup", "section", "report", ".json", ".xml", ".txt"
        )):
            continue
        try:
            raw = zf.read(name)
            text = raw.decode("utf-8", errors="ignore")
        except Exception:
            continue
        lowered = text.lower()
        if "select " not in lowered or " from " not in lowered:
            continue
        for line in text.splitlines():
            lline = line.lower()
            if "select " in lline and " from " in lline:
                stripped = line.strip()
                if stripped:
                    queries.append(stripped)
                    if len(queries) >= 40:
                        break
        if len(queries) >= 40:
            break
    if not queries:
        return ""
    parts: List[str] = ["## SQL-like queries (best effort)"]
    for idx, q in enumerate(queries, start=1):
        shortened = textwrap.shorten(q, width=800, placeholder=" …")
        parts.append(
            f"\n### Query {idx}\n"
            "```sql\n"
            f"{shortened}\n"
            "```"
        )
    return "\n".join(parts).strip()